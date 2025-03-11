import datetime
import os
import threading
import tkinter as tk
from tkinter import messagebox, ttk
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def analisar_posicao():
    arquivo_entrada = "ORIGINAL.xlsx"
    if not os.path.exists(arquivo_entrada):
        messagebox.showerror("Erro", "Arquivo ORIGINAL.xlsx nÃ£o encontrado!")
        return

    # Ler as abas da planilha
    df = pd.read_excel(arquivo_entrada, sheet_name="POSIÃ‡ÃƒO DO DIA", header=0, dtype=str)
    df_clifor = pd.read_excel(arquivo_entrada, sheet_name="CONTROLE DE LIBERAÃ‡ÃƒO - BANNERS", header=None, dtype=str)
    df_base_cliente = pd.read_excel(arquivo_entrada, sheet_name="BASE CLIENTE", header=0, dtype=str)

    df.columns = df.columns.str.strip()

    # AtualizaÃ§Ã£o dos Ã­ndices das colunas
    colunas_posicoes = {
        "VENDEDOR": None,
        "CLIFOR": 0,
        "CLIENTE ATACADO": 1,
        "MATRIZ": None,
        "TIPO": 2,
        "DATA ENVIO": 3,
        "DOCUMENTO": 4,
        "PEDIDO": 8,
        "GRUPO PRODUTO": 9,
        "PRODUTO": 10,
        "DESCRIÃ‡ÃƒO": 11,
        "COR": 12,
        "QTD": 13,
        "VALOR": 14,
        "EMISSÃƒO": 15,
    }

    colunas_reais = {chave: df.columns[posicao] for chave, posicao in colunas_posicoes.items() if posicao is not None}

    # Preencher VENDEDOR e MATRIZ com base na aba "BASE CLIENTE"
    df_base_cliente.columns = df_base_cliente.columns.str.strip()
    mapa_base_cliente = df_base_cliente.set_index(df_base_cliente.columns[0])[[df_base_cliente.columns[1], df_base_cliente.columns[2]]].to_dict(orient="index")

    df["VENDEDOR"] = df["CLIFOR"].map(lambda x: mapa_base_cliente.get(x, {}).get(df_base_cliente.columns[1], ""))
    df["MATRIZ"] = df["CLIFOR"].map(lambda x: mapa_base_cliente.get(x, {}).get(df_base_cliente.columns[2], ""))

    df["DUPLICIDADE"] = df.duplicated(subset=[colunas_reais["CLIFOR"], colunas_reais["PRODUTO"], colunas_reais["COR"]], keep=False)
    df["DUPLICIDADE"] = df["DUPLICIDADE"].apply(lambda x: "Duplicidade" if x else "")

    df["OCORRENCIA"] = df.apply(
        lambda row: (
            "DivergÃªncia de quantidade Footwear"
            if row[colunas_reais["GRUPO PRODUTO"]] == "Footwear" and row[colunas_reais["QTD"]] != "12"
            else "DivergÃªncia de quantidade Apparel"
            if row[colunas_reais["GRUPO PRODUTO"]] == "Apparel" and row[colunas_reais["QTD"]] != "6"
            else "DivergÃªncia de quantidade Accessories"
            if row[colunas_reais["GRUPO PRODUTO"]] == "Accessories" and row[colunas_reais["QTD"]] not in ["3", "4", "5", "6"]
            else ""
        ),
        axis=1
    )

    df[colunas_reais["VALOR"]] = df[colunas_reais["VALOR"]].str.replace(".", ",", regex=False)

    df_filtrado = df[(df["OCORRENCIA"] != "") | (df["DUPLICIDADE"] != "")]
    df_filtrado = df_filtrado[["OCORRENCIA", "DUPLICIDADE"] + list(colunas_reais.values())]
    df_filtrado.columns = ["OCORRENCIA", "DUPLICIDADE"] + list(colunas_reais.keys())

    data_hoje = datetime.datetime.now().strftime("%Y-%m-%d")
    pasta_execucao = os.getcwd()
    arquivo_saida = os.path.join(pasta_execucao, f"ANALISE_POSIÃ‡ÃƒO_{data_hoje}.xlsx")

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        df_filtrado.to_excel(writer, sheet_name="DIVERGENCIA DUPLICIDADE + GRADE", index=False)

    wb = load_workbook(arquivo_saida)
    ws = wb.active
    ws.title = "DIVERGENCIA DUPLICIDADE + GRADE"

    # Filtrar CLIFORs bloqueados
    clifors_bloqueados = df_clifor[df_clifor[5] == "NÃƒO"][0].tolist()
    colunas_para_aba_2 = list(colunas_posicoes.keys())

    df_clifor_nao_pode = df[df[colunas_reais["CLIFOR"]].isin(clifors_bloqueados)]
    df_clifor_nao_pode = df_clifor_nao_pode[[colunas_reais[col] for col in colunas_para_aba_2 if col in colunas_reais]]

    # Criar a segunda aba com o nome correto
    ws2 = wb.create_sheet("CLIFORS QUE NÃƒO PODEM FATURAR")
    for col_idx, col_name in enumerate(df_clifor_nao_pode.columns, 1):
        cell = ws2.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for r_idx, row in enumerate(df_clifor_nao_pode.itertuples(index=False), 2):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if c_idx in [4, 15]:  # Colunas de datas
                cell.number_format = "DD/MM/YYYY"
            elif c_idx == 14:  # Coluna VALOR
                cell.number_format = "R$ #,##0.00"
            elif c_idx in [2, 7]:
                cell.number_format = "000000"

    for ws_name in ["DIVERGENCIA DUPLICIDADE + GRADE", "CLIFORS QUE NÃƒO PODEM FATURAR"]:
        ws_curr = wb[ws_name]
        for col in ws_curr.columns:
            max_length = max(len(str(cell.value)) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_curr.column_dimensions[col_letter].width = max_length + 2

    for col in ["A", "B"]:
        cell = ws[f"{col}1"]
        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        cell.font = Font(bold=True)

    wb.save(arquivo_saida)
    messagebox.showinfo("Sucesso", f"Analise concluída! Planilha salva como {arquivo_saida}")


def realizar_15_30_60_puma():
    arquivo_entrada = "ORIGINAL.xlsx"
    if not os.path.exists(arquivo_entrada):
        messagebox.showerror("Erro", "Arquivo ORIGINAL.xlsx nÃ£o encontrado!")
        return

    try:
        # Carregar as abas
        df_posicao = pd.read_excel(arquivo_entrada, sheet_name="POSIÃ‡ÃƒO 15.30.60", dtype=str)
        df_base_cliente = pd.read_excel(arquivo_entrada, sheet_name="BASE CLIENTE", dtype=str)

        # Ajustar nomes das colunas para evitar erros de espaÃ§os extras
        df_posicao.columns = df_posicao.columns.str.strip().str.upper()
        df_base_cliente.columns = df_base_cliente.columns.str.strip().str.upper()

        # Garantir que CLIFOR seja string para evitar erros na busca
        df_posicao["CLIFOR"] = df_posicao["CLIFOR"].astype(str)
        df_base_cliente["CLIFOR"] = df_base_cliente["CLIFOR"].astype(str)

        # Criar dicionÃ¡rio de busca com base no CLIFOR na BASE CLIENTE
        mapa_clientes = df_base_cliente.set_index("CLIFOR")[["MATRIZ", "VENDEDOR", "CNPJ"]].to_dict("index")

        # Filtrar apenas os pedidos do tipo "03-FATURADO"
        df_filtrado = df_posicao[df_posicao["TIPO"] == "03-FATURADO"].copy()

        # Processar datas
        coluna_data = df_filtrado.columns[15]  # PosiÃ§Ã£o da coluna de data
        df_filtrado["DATA_POS"] = pd.to_datetime(df_filtrado[coluna_data], errors="coerce", format="%Y-%m-%d %H:%M:%S")
        df_filtrado = df_filtrado.dropna(subset=["DATA_POS"])

        hoje = pd.to_datetime(datetime.datetime.today().strftime("%Y-%m-%d"))
        df_filtrado["DIFERENCA"] = (hoje - df_filtrado["DATA_POS"]).dt.days

        # Separar os grupos de 15, 30 e 60 dias
        df_15 = df_filtrado[df_filtrado["DIFERENCA"] == 15].copy()
        df_30 = df_filtrado[df_filtrado["DIFERENCA"] == 30].copy()
        df_60 = df_filtrado[df_filtrado["DIFERENCA"] == 60].copy()

        # Definir colunas a serem usadas no relatÃ³rio final
        colunas_selecionadas = [
            "CLIFOR", "CLIENTE_ATACADO", "MATRIZ", "VENDEDOR", "CNPJ", "TIPO", "DOCUMENTO",
            df_filtrado.columns[15],  # EmissÃ£o
            df_filtrado.columns[13],  # QTDE_R
            df_filtrado.columns[14],  # VALOR_R
        ]

        # FunÃ§Ã£o para processar DataFrames e adicionar informaÃ§Ãµes da BASE CLIENTE
        def processa_df(df):
            if df.empty:
                return pd.DataFrame(columns=colunas_selecionadas)

            df_sel = df[["CLIFOR", "CLIENTE_ATACADO", "TIPO", "DOCUMENTO", df_filtrado.columns[15],
                         df_filtrado.columns[13], df_filtrado.columns[14]]].copy()

            # Adicionar as informaÃ§Ãµes da BASE CLIENTE
            df_sel["MATRIZ"] = df_sel["CLIFOR"].map(lambda x: mapa_clientes.get(x, {}).get("MATRIZ", ""))
            df_sel["VENDEDOR"] = df_sel["CLIFOR"].map(lambda x: mapa_clientes.get(x, {}).get("VENDEDOR", ""))
            df_sel["CNPJ"] = df_sel["CLIFOR"].map(lambda x: mapa_clientes.get(x, {}).get("CNPJ", ""))

            # Ajustar formataÃ§Ãµes de data e nÃºmeros
            df_sel[df_sel.columns[4]] = pd.to_datetime(df_sel[df_sel.columns[4]], errors="coerce").dt.strftime("%d/%m/%Y")

            df_sel[df_sel.columns[5]] = pd.to_numeric(
                df_sel[df_sel.columns[5]].str.replace(",", ".").str.replace("[^0-9.]", "", regex=True), errors="coerce"
            ).fillna(0)
            df_sel[df_sel.columns[6]] = pd.to_numeric(
                df_sel[df_sel.columns[6]].str.replace(",", ".").str.replace("[^0-9.]", "", regex=True), errors="coerce"
            ).fillna(0)

            # Agrupar dados por DOCUMENTO
            df_grouped = df_sel.groupby("DOCUMENTO", as_index=False).agg({
                "CLIFOR": "first",
                "CLIENTE_ATACADO": "first",
                "MATRIZ": "first",
                "VENDEDOR": "first",
                "CNPJ": "first",
                "TIPO": "first",
                "DOCUMENTO": "first",
                df_sel.columns[4]: "first",  # EmissÃ£o
                df_sel.columns[5]: "sum",    # QTDE_R
                df_sel.columns[6]: "sum",    # VALOR_R
            })

            return df_grouped

        # Processar os DataFrames filtrados
        df_15_final = processa_df(df_15)
        df_30_final = processa_df(df_30)
        df_60_final = processa_df(df_60)

        # Gerar arquivo Excel com os relatÃ³rios
        data_hoje = datetime.datetime.today().strftime("%Y-%m-%d")
        arquivo_saida = f"15.30.60_Puma_{data_hoje}.xlsx"
        with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
            df_15_final.to_excel(writer, sheet_name="15", index=False)
            df_30_final.to_excel(writer, sheet_name="30", index=False)
            df_60_final.to_excel(writer, sheet_name="60", index=False)

        # Ajustar largura das colunas no Excel gerado
        wb = load_workbook(arquivo_saida)
        for aba in ["15", "30", "60"]:
            ws = wb[aba]
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[column].width = max_length + 2

        wb.save(arquivo_saida)
        messagebox.showinfo("Sucesso", f"Arquivo {arquivo_saida} gerado com sucesso!")

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro ao gerar o arquivo 15/30/60 PUMA: {e}")


def analisar_carteira():
    def task():
        botao_analise_carteira.config(state="disabled")
        progress_bar.pack(pady=20)
        progress_bar.start()

        arquivo_entrada = "ORIGINAL.xlsx"
        if not os.path.exists(arquivo_entrada):
            messagebox.showerror("Erro", "Arquivo ORIGINAL.xlsx não encontrado!")
            return

        try:
            # Carregar as planilhas
            df_atual = pd.read_excel(
                arquivo_entrada, sheet_name="CARTEIRA ATUAL", header=0, dtype=str
            )
            
            df_cancelados = pd.read_excel(
                arquivo_entrada, sheet_name="CONTROLE DE CANCELAMENTOS", header=0, dtype=str
            )
            df_alteracoes = pd.read_excel(
                arquivo_entrada, sheet_name="ALTERAÇÃO DA DATA DE ENTREGA", header=0, dtype=str
            )

            # Normalizar os cabeÃ§alhos para evitar erros de espaço ou maiÃºsculas
            df_atual.columns = df_atual.columns.str.strip().str.upper()
            df_cancelados.columns = (
                df_cancelados.columns.str.strip()
                .str.upper()
                .str.normalize("NFKD")
                .str.encode("ascii", errors="ignore")
                .str.decode("utf-8")
            )
            df_alteracoes.columns = (
                df_alteracoes.columns.str.strip()
                .str.upper()
                .str.normalize("NFKD")
                .str.encode("ascii", errors="ignore")
                .str.decode("utf-8")
            )

            # Verificar se 'CANCELAMENTO?' e 'PEDIDO' estão presentes
            if (
                df_cancelados.shape[1] < 2
                or df_alteracoes.shape[1] < 2
            ):
                messagebox.showerror(
                    "Erro",
                    f"Dados de 'CANCELAMENTO' ou 'PEDIDO' nÃ£o encontradas em {arquivo_entrada}.",
                )
                return

            # Filtrar pedidos para cancelar
            pedidos_cancelados = df_cancelados["PEDIDO"].str.cat(
                [
                    df_cancelados["PRODUTO"], 
                    df_cancelados["COR"].apply(lambda x: str(int(x)) if pd.notnull(x) else x)
                 ], sep="-"
            ).unique()

            # Verificar se os pedidos cancelados ainda estão na carteira atual
            pedidos_na_carteira = df_atual[
                        df_atual["PEDIDO"].str.cat([
                            df_atual["PRODUTO"], 
                            df_atual["COR"].apply(lambda x: str(int(x)) if pd.notnull(x) else x)], sep="-")
                            .isin(pedidos_cancelados)
                            ]

            if pedidos_na_carteira.empty:
                messagebox.showinfo(
                    "Resultado", "Nenhum pedido cancelado encontrado na carteira atual!"
                )
                return

            # Selecionar as colunas conforme os í­ndices informados
            colunas_indices = {
                "CLIFOR": df_atual.columns[31],  # Coluna A
                "PEDIDO": df_atual.columns[0],  # Coluna B
                "PRODUTO": df_atual.columns[8],  # Coluna AG
                "COR": df_atual.columns[35],  # Coluna AI
                "ENTREGA": df_atual.columns[13],  # Coluna R (QTD ALOCADA)
                "ITEM": df_atual.columns[7],  # Coluna AA
            }
            # Salvar a planilha de análise dos cancelados
            data_hoje = datetime.datetime.now().strftime("%Y-%m-%d")
            pasta_execucao = os.getcwd()
            arquivo_saida = os.path.join(
                pasta_execucao, f"ANALISE_CARTEIRA_{data_hoje}.xlsx"
            )
            df_cancelados = pedidos_na_carteira[list(colunas_indices.values())]
            df_cancelados.columns = colunas_indices.keys()
            ordem_colunas = [
                "CLIFOR",
                "PEDIDO",
                "PRODUTO",
                "COR",
                "ENTREGA",
                "ITEM",
            ]
            df_cancelados = df_cancelados[ordem_colunas]
            with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
                df_cancelados.to_excel(
                    writer, sheet_name="CANCELADOS AINDA EM CARTEIRA", index=False
                )

            # Ajustar layout da planilha dos cancelados
            wb = load_workbook(arquivo_saida)
            ws = wb.active
            # Formatar coluna VALOR TOTAL como moeda
            for row in ws.iter_rows(
                min_row=2, min_col=7, max_col=7, max_row=ws.max_row
            ):
                for cell in row:
                    try:
                        valor = float(cell.value.replace(",", "").replace(".", "."))
                        cell.value = valor
                        cell.number_format = '"R$" #,##0.00_-'
                    except ValueError:
                        pass
                    except AttributeError:
                        pass
            # Ajuste automÃ¡tico de largura
            for column_cells in ws.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[col_letter].width = max_length + 2
            wb.save(arquivo_saida)

        except Exception as e:
            progress_bar.stop()
            progress_bar.pack_forget()
            messagebox.showerror(
                "Erro", f"Ocorreu um erro durante a análise 1 da carteira: {e}"
            )

        # ================================
        # Nova funcionalidade: Alterações de Datas Não Refletidas (sem mudanças)
        # ================================
        # TODO - Alteração de datas agora é uma aba ALTERAÇÃO DA DATA DE ENTREGA
        try:
            col_entrega_antiga = df_alteracoes.columns[3]  # Coluna E
            col_entrega_correta = df_alteracoes.columns[5]  # Coluna F
            df_acoes_filtrado = df_alteracoes[
                (df_alteracoes[col_entrega_antiga].notna())
                & (df_alteracoes[col_entrega_antiga] != "")
                & (df_alteracoes[col_entrega_correta].notna())
                & (df_alteracoes[col_entrega_correta] != "")
            ]

            df_merged = pd.merge(
                df_atual,
                df_acoes_filtrado,
                on="PEDIDO",
                how="inner",
                suffixes=("", "_acoes"),
            )

            # TODO - os nomes vão mudar
            col_entrega_atual = df_atual.columns[13]
            col_entrega_correta = df_atual.columns[13] + "_acoes"

            df_merged["DATA_ENTREGA_CORRETA"] = pd.to_datetime(
                df_merged[col_entrega_correta], errors="coerce", dayfirst=True
            )
            df_merged["DATA_ENTREGA_ATUAL"] = pd.to_datetime(
                df_merged[col_entrega_atual], errors="coerce", dayfirst=True
            )
            df_alteracoes = df_merged[
                df_merged["DATA_ENTREGA_CORRETA"] != df_merged["DATA_ENTREGA_ATUAL"]
            ].copy()

            df_alteracoes_final = pd.DataFrame(
                {
                    "PEDIDO": df_alteracoes["PEDIDO"],
                    "PRODUTO": df_alteracoes["PRODUTO"],
                    "COR": df_alteracoes["COR"],
                    "DATA ATUAL": df_alteracoes[col_entrega_atual],
                    "ITEM": df_alteracoes["ITEM"],
                    "ENTREGA": df_alteracoes[col_entrega_correta],
                    "LIMITE ENTREGA": df_alteracoes["LIMITE DE ENTREGA"],
                    "DATA ATUAL DA CARTEIRA": df_alteracoes[df_atual.columns[13]],

                    # TODO - nova coluna DATA ENTREGA ATUAL (data errada)
                }
            )
            with pd.ExcelWriter(arquivo_saida, engine="openpyxl", mode="a") as writer:
                df_alteracoes_final.to_excel(
                    writer, sheet_name="ALTERAÇÕES DE DATAS N REFL", index=False
                )
        except Exception as e:
            progress_bar.stop()
            progress_bar.pack_forget()
            messagebox.showerror(
                "Erro", f"Ocorreu um erro na análise de alterações de datas: {e}"
            )

        # Ajustar layout e formataÃ§Ã£o na aba "ALTERAÃ‡Ã•ES DE DATAS N REFL"
        try:
            wb = load_workbook(arquivo_saida)
            ws_alt = wb["ALTERAÇÕES DE DATAS N REFL"]
            for column_cells in ws_alt.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                ws_alt.column_dimensions[col_letter].width = max_length + 2
            from datetime import datetime as dt

            for col in ["D", "E"]:
                for cell in ws_alt[col]:
                    if cell.row > 1 and isinstance(cell.value, str):
                        try:
                            cell.value = dt.strptime(cell.value.split()[0], "%Y-%m-%d")
                        except:
                            pass
                    cell.number_format = "dd/mm/yyyy"
            ws_alt["H1"].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            wb.save(arquivo_saida)
        except Exception as e:
            progress_bar.stop()
            progress_bar.pack_forget()
            messagebox.showerror(
                "Erro",
                f"Ocorreu um erro ao ajustar a formatação da aba ALTERAÇÕES DE DATAS N REFL: {e}",
            )

        # ================================
        try:
            # Ler as abas CARTEIRA ANTERIOR e CARTEIRA ATUAL
            df_anterior = pd.read_excel(
                arquivo_entrada, sheet_name="CARTEIRA ANTERIOR", header=0, dtype=str
            )
            df_anterior.columns = df_anterior.columns.str.strip().str.upper()
            df_atual = pd.read_excel(
                arquivo_entrada, sheet_name="CARTEIRA ATUAL", header=0, dtype=str
            )
            df_atual.columns = df_atual.columns.str.strip().str.upper()
            # Definir os Ã­ndices fixos conforme solicitado
            idx_pedido = 0  # PEDIDO (coluna A)
            idx_cliente = 1  # CLIENTE ATACADO (coluna B)
            idx_qtd = 17  # QTD (coluna R)
            idx_matriz = 32  # MATRIZ (da CARTEIRA ATUAL)
            idx_data = 57  # DATA DE ALOCAÃ‡ÃƒO (coluna BF)
            idx_valor = 22  # VALOR (coluna W)
            idx_clifor = 31  # CLIFOR (coluna AF) â€“ somente da CARTEIRA ATUAL

            # Obter os nomes das colunas
            col_pedido = df_anterior.columns[idx_pedido]
            col_cliente_ant = df_anterior.columns[idx_cliente]
            col_cliente_atu = df_atual.columns[idx_cliente]
            col_qtd_ant = df_anterior.columns[idx_qtd]
            col_qtd_atu = df_atual.columns[idx_qtd]
            col_matriz_atu = df_atual.columns[idx_matriz]
            col_data_ant = df_anterior.columns[idx_data]
            col_data_atu = df_atual.columns[idx_data]
            col_valor_ant = df_anterior.columns[idx_valor]
            col_valor_atu = df_atual.columns[idx_valor]
            col_clifor = df_atual.columns[idx_clifor]  # referencia a coluna CLIFOR

            # Realizar merge utilizando PEDIDO como chave
            df_merge = pd.merge(
                df_anterior,
                df_atual,
                on=col_pedido,
                how="inner",
                suffixes=("_ant", "_atu"),
            )
            cond = (
                (
                    (df_merge[col_qtd_ant + "_ant"] != df_merge[col_qtd_atu + "_atu"])
                    | (
                        df_merge[col_data_ant + "_ant"]
                        != df_merge[col_data_atu + "_atu"]
                    )
                )
                & (df_merge[col_data_ant + "_ant"].notna())
                & (df_merge[col_data_ant + "_ant"] != "")
                & (df_merge[col_data_atu + "_atu"].notna())
                & (df_merge[col_data_atu + "_atu"] != "")
            )
            df_diff = df_merge[cond].copy()
            if col_clifor in df_diff.columns:
                col_clifor_final = col_clifor
            elif (col_clifor + "_atu") in df_diff.columns:
                col_clifor_final = col_clifor + "_atu"
            else:
                col_clifor_final = col_clifor
            df_comp_final = pd.DataFrame(
                {
                    "PEDIDO": df_diff[col_pedido],
                    "CLIENTE ATACADO ATUAL": df_diff[col_cliente_atu + "_atu"],
                    "CLIENTE ATACADO ANTERIOR": df_diff[col_cliente_ant + "_ant"],
                    "CLIFOR": df_diff[col_clifor_final],
                    "MATRIZ": df_diff[col_matriz_atu + "_atu"],
                    "QTD ANTERIOR": df_diff[col_qtd_ant + "_ant"],
                    "QTD ATUAL": df_diff[col_qtd_atu + "_atu"],
                    "DATA ALOCAÇÃO ANTERIOR": df_diff[col_data_ant + "_ant"],
                    "DATA ALOCAÇÃO ATUAL": df_diff[col_data_atu + "_atu"],
                    "VALOR ATUAL": df_diff[col_valor_atu + "_atu"],
                    "VALOR ANTERIOR": df_diff[col_valor_ant + "_ant"],
                }
            )

            def stts(row):
                alterations = []
                if row["QTD ANTERIOR"] != row["QTD ATUAL"]:
                    alterations.append("QTD ALTERADA")
                if row["DATA ALOCAÇÃO ATUAL"] != row["DATA ALOCAÇÃO ANTERIOR"]:
                    alterations.append("DATA DE CHEGADA ALTERADA")
                return " + ".join(alterations)

            df_comp_final.insert(0, "STTS", df_comp_final.apply(stts, axis=1))
            with pd.ExcelWriter(arquivo_saida, engine="openpyxl", mode="a") as writer:
                df_comp_final.to_excel(
                    writer, sheet_name="COMPARATIVO DE ALOCAÇÃO", index=False
                )
        except Exception as e:
            progress_bar.stop()
            progress_bar.pack_forget()
            messagebox.showerror(
                "Erro", f"Ocorreu um erro ao criar o comparativo de ALOCAÇÃO: {e}"
            )

        # ================================
        # Formatar a aba "COMPARATIVO DE ALOCAÇÃO"
        # ================================
        try:
            wb = load_workbook(arquivo_saida)
            ws_comp = wb["COMPARATIVO DE ALOCAÇÃO"]
            for column_cells in ws_comp.columns:
                max_length = 0
                col_letter = get_column_letter(column_cells[0].column)
                for cell in column_cells:
                    if cell.value is not None:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                ws_comp.column_dimensions[col_letter].width = max_length + 2
            ws_comp["A1"].fill = PatternFill(
                start_color="FFFF00", end_color="FFFF00", fill_type="solid"
            )
            from datetime import datetime as dt

            for row in ws_comp.iter_rows(
                min_row=2, min_col=9, max_col=10, max_row=ws_comp.max_row
            ):
                for cell in row:
                    try:
                        if isinstance(cell.value, str):
                            cell.value = dt.strptime(cell.value.split()[0], "%Y-%m-%d")
                        cell.number_format = "dd/mm/yyyy"
                    except Exception:
                        pass
            for row in ws_comp.iter_rows(
                min_row=2, min_col=11, max_col=12, max_row=ws_comp.max_row
            ):
                for cell in row:
                    try:
                        cell.value = float(cell.value)
                        cell.number_format = '"R$"#,##0.00'
                    except Exception:
                        pass
            wb.save(arquivo_saida)
        except Exception as e:
            messagebox.showerror(
                "Erro", f"Ocorreu um erro ao formatar o comparativo de alocação: {e}"
            )
        progress_bar.stop()
        progress_bar.pack_forget()
        botao_analise_carteira.config(state="normal")

        messagebox.showinfo(
            "Sucesso",
            f"Analise de carteira concluída! Planilha salva como {arquivo_saida}",
        )

    threading.Thread(target=task, daemon=True).start()

# ConfiguraÃ§Ã£o da interface Tkinter
janela = tk.Tk()
janela.title("Automações Escritório Ativação")
janela.geometry("500x350")

titulo = tk.Label(janela, text="Automações", font=("Century Gothic", 12))
titulo.pack(pady=10)

progress_bar = ttk.Progressbar(
    janela, orient="horizontal", mode="determinate", length=300
)

botao_analisar = tk.Button(
    janela, text="Analisar PosiÃ§Ã£o", command=analisar_posicao, font=("Century Gothic", 12,"bold")
)
botao_analisar.pack(pady=15)


botao_15_30_60_puma = tk.Button(
    janela,
    text="Realizar 15/30/60 PUMA",
    command=realizar_15_30_60_puma,
    font=("Century Gothic", 12,"bold"),
)
botao_15_30_60_puma.pack(pady=15)

botao_analise_carteira = tk.Button(
    janela, text="Analisar Carteira", command=analisar_carteira, font=("Century Gothic", 12,"bold")
)
botao_analise_carteira.pack(pady=15)

janela.mainloop()